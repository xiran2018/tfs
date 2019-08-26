#coding:utf-8

import argparse
import xlrd


from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import Color, Fill, Font, Alignment, PatternFill, Border, Side
from openpyxl.cell import Cell
import datetime
import string

import os, sys
import random

# 单个文件执行命令
#python main.py -n "D:/PythonWorkSpace/tfs2excel/1. 发货清单，贾海侠，2019.04.03.xlsx" -s 2019年销量统计 -l 242
# 多个文件执行命令
#python main.py -n "D:/PythonWorkSpace/tfs2excel/1. 发货清单，贾海侠，2019.04.03.xlsx" -s 2019年销量统计 -l 242#241#23

parser = argparse.ArgumentParser(description='生成tfs公司的报价表需要的参数')
# General options


parser.add_argument('-n','--excel_file_name',required = True, type=str,
                    help='name of the excel(include the absoulate location). \
                        It decides where to find the file')
parser.add_argument('-s','--sheet_name', required = True,default='', type=str,
                    help='name of the sheet for the excel file.')

parser.add_argument('-l','--lines',required = True,type=str, help='line number for generate excel,\
    it can split by #,such as 15#20#30')

parser.add_argument('-o','--generate_Location',required = False,default='',type=str,
    help='location of the generate excel file')


def isdigit(aString):
    try:
        x = float(aString)
        return True
    except ValueError as e:
        # print('input is not a number!')
        return False


def readLines(excelName,sheetName,linNumber):
    #打开excel文件
    data=xlrd.open_workbook(excelName)
    #获取第一张工作表（通过索引的方式）
    table = data.sheet_by_name(sheetName)
    #data_list用来存放数据
    data_list=[]
    #将table中第一行的数据读取并添加到data_list中
    data_list.extend(table.row_values(linNumber))

    return data_list


def generateExcel(content):

    #记不记抛
    # if 重量/体积<125
    #     重量=125*体积
    # else
    #     重量=重量

    # 创建一个工作薄
    wb = Workbook()

    # 创建一个工作表(注意是一个属性) # 激活 worksheet
    table = wb.active

    # excel创建的工作表名默认为sheet1,一下代码实现了给新创建的工作表创建一个新的名字
    table.title = '托运单'

    # 合并C1 D1
    # 法一
    # table.merge_cells('C1:D1')
    # table.cell(row = 1,column = 3,value = 'pdf/mp3链接')
    # 法二

    # 设置表头字体居中
    font = Font(name=u'宋体',size=16, bold = True)
    alignment_style = Alignment(horizontal='center', vertical='center')

    # 定义Border边框样式
    left, right, top, bottom = [Side(style='thin', color='0000ff')]*4
    border_style = Border(left=left, right=right, top=top, bottom=bottom)

    #第一行
    table.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    table.cell(row=1, column=1,value = '天府盛（TFS）俄罗斯海外仓货物托运单')
    table.cell(row=1, column=1).font = font
    # 水平对齐 垂直对齐
    table.cell(row=1, column=1).alignment = alignment_style
    table.cell(row=1, column=1).border = border_style

    #第二行
    table.cell(row=2, column=1,value = '运输方式')
    table.cell(row=2, column=2, value=content[4])
    table.cell(row=2, column=3,value = '收货仓库')
    table.cell(row=2, column=4, value=content[3])
    table.cell(row=2, column=5,value = '签出日期')

    # 日期
    now_time = datetime.datetime.now().strftime('%Y%m%d %H%M')
    # 格式化时间为中文年月日
    # now_time = now_time[:4] + '年' + now_time[4:6] + '月' + now_time[6:8] + '日' + now_time[8:11] + '时' + now_time[11:13] + '分'
    now_time = now_time[:4] + '年' + now_time[4:6] + '月' + now_time[6:8] + '日'
    table.cell(row=2, column=6, value=now_time)

    # 第三行
    table.cell(row=3, column=1, value='委托方')
    table.cell(row=3, column=1).font = font
    table.merge_cells(start_row=3, start_column=2, end_row=3, end_column=6)
    table.cell(row=3, column=2, value=content[5])

    # 第四行
    table.cell(row=4, column=1, value='承运方')
    table.cell(row=4, column=1).font = font
    table.merge_cells(start_row=4, start_column=2, end_row=4, end_column=6)
    table.cell(row=4, column=2, value='天府盛（北京）国际供应链管理有限公司')

    #第五行
    table.cell(row=5, column=1,value = '客户编码')
    table.cell(row=5, column=2, value=content[7])
    table.cell(row=5, column=3,value = '客户单证号')
    table.cell(row=5, column=4, value=content[9])
    table.cell(row=5, column=5,value = '票号')
    table.cell(row=5, column=6, value=content[10])

    #第6行
    table.cell(row=6, column=1,value = '品种')
    table.cell(row=6, column=2, value=content[11])
    table.cell(row=6, column=3,value = '总件数')
    tempStr=content[10].split("-")
    table.cell(row=6, column=4, value=tempStr[1])
    table.cell(row=6, column=5,value = '单品数量(个)')
    table.cell(row=6, column=6, value=content[12])

    #第7行
    table.cell(row=7, column=1,value = '体积单价(￥/m³)')
    table.cell(row=7, column=2, value='0')
    table.cell(row=7, column=3,value = '体积(m³)')
    table.cell(row=7, column=4, value=content[15])
    table.cell(row=7, column=5,value = '体积运费(￥)')
    table.cell(row=7, column=6, value='0')

    #第8行
    table.cell(row=8, column=1,value = '重量单价(￥/Kg)')
    table.cell(row=8, column=2, value=content[26])
    table.cell(row=8, column=3,value = '重量(Kg)')
    table.cell(row=8, column=4, value=content[14])
    table.cell(row=8, column=5,value = '重量运费(￥)')
    table.cell(row=8, column=6, value=content[28])

    #第9行
    table.cell(row=9, column=1,value = '保险金额(￥/Kg)')
    price = content[13]
    if(isdigit(price)):
        table.cell(row=9, column=2, value=price)
    else:
        table.cell(row=9, column=2, value='0')
    table.cell(row=9, column=3,value = '保险率')
    table.cell(row=9, column=4, value=content[27])
    table.cell(row=9, column=5,value = '保险费(￥)')
    table.cell(row=9, column=6, value=content[30])

    #第10行
    table.cell(row=10, column=1,value = '保险金额(￥)')
    table.cell(row=10, column=2, value=content[29])
    table.cell(row=10, column=3,value = '叉车费')
    table.cell(row=10, column=4, value=content[32])
    table.cell(row=10, column=5,value = '总费用')
    firstTotalPrice=int(float(content[31])+0.5)
    table.cell(row=10, column=6, value=firstTotalPrice)

    font10 = Font(name=u'宋体', size=11, bold=True)
    table.cell(row=10, column=5).font = font10
    table.cell(row=10, column=6).font = font10

    #第12行
    table.merge_cells(start_row=12, start_column=1, end_row=12, end_column=6)
    table.cell(row=12, column=1, value='请24小时内核实运单信息，过期将视为默认，一旦入账信息不再更改。谢谢配合！')
    font12 = Font(name=u'宋体', color="FF0000",size=11, bold=False)
    table.cell(row=12, column=1).font = font12
    # 水平对齐 垂直对齐
    table.cell(row=12, column=1).alignment = alignment_style
    table.cell(row=12, column=1).border = border_style

    # 第14行
    table.cell(row=14, column=1, value='一、企业支付宝账号：')
    font14 = Font(name=u'宋体', size=11, bold=True)
    table.cell(row=14, column=1).font = font14

    # 第15行
    table.cell(row=15, column=1, value='136685584@qq.com')
    # font14 = Font(name=u'宋体', size=11, bold=True)
    # table.cell(row=14, column=1).font = font14

    # 第16行
    table.cell(row=16, column=1, value='用户名称：天府盛（北京）国际供应链管理有限公司。')

    # 第17行
    table.cell(row=17, column=1, value='二、公司银行账户')
    table.cell(row=17, column=1).font = font14

    # 第18行
    table.cell(row=18, column=1, value='公司名称：天府盛（北京）国际供应链管理有限公司')

    # 第19行
    table.cell(row=19, column=1, value='开户行：中信银行北京东大桥支行')

    # 第20行
    table.cell(row=20, column=1, value='账号：8110701013900022960')

    applyStyle(table, "B5:B10", alignment_style)
    applyStyle(table, "D5:D10", alignment_style)
    applyStyle(table, "F5:F10", alignment_style)




    # 调整列宽
    # table.column_dimensions['A'].width = 20.0

    # 生成前14个大写字母  ascii_uppercase生成所有大写字母
    upper_string = string.ascii_uppercase[:6]
    for col in upper_string:
        table.column_dimensions[col].width = 15

    # 调整行高
    table.row_dimensions[1].height = 35

    fileName="天府盛（TFS）俄罗斯海外仓货物托运单_{0}_{1}元.xlsx".format(content[9],firstTotalPrice)
    if(delFile(fileName)):
        wb.save(fileName)
        print("保存文件成功:{}".format(fileName))


def delFile(file):

    try:
        # 判断文件是否存在
        if (os.path.exists(file)):
            os.remove(file)
            # print('移除文件：%s' % file)
        # else:
        #     print("无相同文件不用删除！")
        return True
    except Exception as e:
        print('文件{}删除错误，可能文件被打开了，请关闭重试!'.format(file))
        return False




def applyStyle(table,rangeString,alignment_style):

    ##################################全局居中
    commonBackgroundColorHex = "AACF91"

    commonFill = PatternFill(start_color=commonBackgroundColorHex, end_color=commonBackgroundColorHex,
                             fill_type="solid")

    for eachCommonRow in table.iter_rows(rangeString):

        # logging.info("eachCommonRow=%s", eachCommonRow)

        for eachCellInRow in eachCommonRow:
            # logging.info("eachCellInRow=%s", eachCellInRow)

            eachCellInRow.alignment = alignment_style

            # eachCellInRow.fill = commonFill


def main(opts):


    if opts.excel_file_name == '':
        print("please set the file name")
        exit()
    if opts.sheet_name == '':
        print("please set the sheet name")
        exit()
    if opts.lines == '':
        print("please set the lines")
        exit()

    lineStr=opts.lines
    lines=lineStr.split('#')
    print("+++++++++++++++lines:"+str(len(lines)))
    for line in lines:
        realLine = int(line)-1
        content=readLines(opts.excel_file_name,opts.sheet_name,realLine)
        # print("行{0}的列数为{1}".format(line,len(content)))
        # print(content)
        # print("####################")
        # print(content[20])
        generateExcel(content)
        print("行{0}处理完毕\n".format(line))
        # generateExcel()




    print("--> Finished generate excel")


if __name__ == '__main__':
    opts = parser.parse_args()
    main(opts)