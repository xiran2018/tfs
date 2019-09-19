#coding:utf-8

import argparse
import xlrd
from .models.companyInfo import *
sys.path.append("..")
from db import db


from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import Color, Fill, Font, Alignment, PatternFill, Border, Side
from openpyxl.cell import Cell
import datetime
import string

import os, sys
import random

import pymysql

# 执行命令
# python main.py -s 32 -e 49 -f "22"


parser = argparse.ArgumentParser(description='从数据库中excel')
# General options



parser.add_argument('-s','--start', required = False,default=-1, type=int,
                    help='开始行位置。')

parser.add_argument('-e','--end',required = False,default=-1,type=int, help='结束行位置。')

parser.add_argument('-f','--file_name', required = True,default='', type=str, help='name of the  file.')


def getHaveReadInfo():
    fname = 'haveRead.txt'
    with open(fname, 'r', encoding='utf-8') as f:  # 打开文件
        lines = f.readlines()  # 读取所有行
        if(lines != "" and len(lines)>0):
            first_line = lines[0]  # 取第一行
            # last_line = lines[-1]  # 取最后一行
            print('文件' + fname + '第一行为：' + first_line)
            # print('文件' + fname + '最后一行为：' + last_line)
            return first_line
        else:
            return "已读信息"
def isdigit(aString):
    try:
        x = float(aString)
        return True
    except ValueError as e:
        # print('input is not a number!')
        return False

def getUserInfo():

    user={}

    # 打开数据库连接
    db = pymysql.connect(host='13.231.165.68', port=33070, user='root', passwd='880309jQl', db='tfs', charset='utf8')

    # 使用cursor()方法获取操作游标
    cursor = db.cursor()


    sql = "SELECT * FROM  user"

    try:
        # 执行SQL语句
        cursor.execute(sql)
        # 获取所有记录列表
        results = cursor.fetchall()

        for row in results:
            user[row[0]]=row[5]
        # print(len(results))

    except:
        print("Error: unable to fecth user data")

    # 关闭数据库连接
    db.close()
    return user;


def getData(start,companyName,storeName,userId,bianhao,shengfen,city, \
                          countNumber, canbaorenshu, isHaveCang, startNumber, \
                          endNumber, yingxiaocount, startTime,endTime,isFromFileNumber):
    if(isFromFileNumber==1):
        print("===============从哪一行开始读取文件=================={}".format(start))
        # print("=================end=================={}".format(end))
        pn = CompanyInfo.query.filter(
            CompanyInfo.id>start if start is not None and start != -1 else "",
        ).all()
    else:
        pn = CompanyInfo.query.filter(
            CompanyInfo.companyName.like("%" + companyName + "%") if companyName is not None and companyName != "" else "",
            CompanyInfo.storeName.like("%" + storeName + "%") if storeName is not None and storeName != "" else "",
            CompanyInfo.belongTo == userId if userId is not None and userId != "" else "",
            CompanyInfo.customBianHao.like("%" + bianhao + "%") if bianhao is not None and bianhao != "" else "",
            CompanyInfo.shengfen.like("%" + shengfen + "%") if shengfen is not None and shengfen != "" else "",
            CompanyInfo.shengfen.like("%" + city + "%") if city is not None and city != "" else "",
            CompanyInfo.countNumber>= countNumber if countNumber is not None and countNumber != "" else "",
            CompanyInfo.canbaorenshu> canbaorenshu if canbaorenshu is not None and canbaorenshu != "" else "",
            CompanyInfo.isHaveHaiCang==isHaveCang if isHaveCang is not None and isHaveCang != "" else "",
            CompanyInfo.id.between(startNumber,endNumber) if startNumber is not None and startNumber != "" and endNumber is not None and endNumber != "" else "",
            CompanyInfo.id > startTime if startTime is not None and startTime != "" and (endNumber is None or endNumber == "") else "",
            CompanyInfo.cuxiaocount == yingxiaocount if yingxiaocount is not None and yingxiaocount != "" else "",
            CompanyInfo.updatetime.between(startTime,endTime) if startTime is not None and startTime != "" and endTime is not None and endTime != "" else "",
            CompanyInfo.updatetime > startTime if startTime is not None and startTime != "" and (endTime is None or endTime == "") else "",
        ).all()
    # pn = CompanyInfo.query.filter(CompanyInfo.id > 10052).all()
    # print("================长度如下===================")
    # print(len(pn))
    return pn


def readFile(path):
    with open(path, 'r') as my_file:
        # values2 = my_file.readlines()
        # values2 = my_file.read()
        values2 = my_file.readline()
        return values2

def writeFile(path,content):

    with open(path, 'w') as my_file:
        # values2 = my_file.readlines()
        # values2 = my_file.read()
        values2 = my_file.write(content)

def getTelAndMail(TelAndMailInfos):
    tel=""
    dianhua=""
    mail=""
    qq = ""
    for ele in TelAndMailInfos:
        type = ele.type
        value = ele.value
        status = ele.status
        beizhu = ele.beizhu
        if beizhu!=None:
            value = value+"#"+str(status)+"|"+beizhu
        else:
            value = value+"#" + str(status)

        if(type == 1):
            if(tel==""):
                tel = value
            else:
                tel = tel+";"+value
        elif(type==2):
            if (dianhua == ""):
                dianhua = value
            else:
                dianhua = dianhua + ";" + value
        elif(type==3):
            if (qq == ""):
                qq = value
            else:
                qq = qq + ";" + value
        elif (type == 4):
            if (mail == ""):
                mail = value
            else:
                mail = mail + ";" + value

    return tel,dianhua,mail,qq
def generateExcel(results,target_path,excelName,isFromFileNumber):



    # 创建一个工作薄
    wb = Workbook()

    # 创建一个工作表(注意是一个属性) # 激活 worksheet
    table = wb.active

    # excel创建的工作表名默认为sheet1,一下代码实现了给新创建的工作表创建一个新的名字
    table.title = 'ali数据'


    # 设置表头字体居中
    font = Font(name=u'宋体',size=16, bold = True)
    alignment_style = Alignment(horizontal='center', vertical='center')

    # print(type(results))
    i=1
    # nlist = range(1, 25) # 生成  1  到 24
    # for j in nlist:
    table.cell(row=i, column=1, value="序号")
    table.cell(row=i, column=2, value="促销次数")
    table.cell(row=i, column=3, value="所属人员")
    table.cell(row=i, column=4, value="用户编号")
    table.cell(row=i, column=5, value="营销渠道")
    table.cell(row=i, column=6, value="备注")
    table.cell(row=i, column=7, value="店铺数量")
    table.cell(row=i, column=8, value="有/无仓")
    table.cell(row=i, column=9, value="分类")
    table.cell(row=i, column=10, value="公司名称")
    table.cell(row=i, column=11, value="商店名称")
    table.cell(row=i, column=12, value="手机")
    table.cell(row=i, column=13, value="电话")
    table.cell(row=i, column=14, value="QQ")
    table.cell(row=i, column=15, value="邮箱")
    table.cell(row=i, column=16, value="店铺链接(|)")
    table.cell(row=i, column=17, value="网址(|)")

    table.cell(row=i, column=18, value="经营状态")
    table.cell(row=i, column=19, value="法人代表")
    table.cell(row=i, column=20, value="注册资本")
    table.cell(row=i, column=21, value="成立日期")
    table.cell(row=i, column=22, value="所属省份")
    table.cell(row=i, column=23, value="所属城市")
    table.cell(row=i, column=24, value="社会信用码")
    table.cell(row=i, column=25, value="纳税人识别号")
    table.cell(row=i, column=26, value="注册号")
    table.cell(row=i, column=27, value="组织机构代码")
    table.cell(row=i, column=28, value="参保人数")
    table.cell(row=i, column=29, value="企业类型")
    table.cell(row=i, column=30, value="所属行业")
    table.cell(row=i, column=31, value="企业地址")
    table.cell(row=i, column=32, value="经营范围")
    table.cell(row=i, column=33, value="登记机关")
    table.cell(row=i, column=34, value="来源")
    table.cell(row=i, column=35, value="更新时间")

    lastRowNumber=-1
    for row in results:
        lastRowNumber=row.id
        i = i + 1
        table.cell(row=i, column=1, value=row.id)
        table.cell(row=i, column=2, value=row.cuxiaocount)
        table.cell(row=i, column=3, value=row.user.realname)
        table.cell(row=i, column=4, value=row.customBianHao)
        table.cell(row=i, column=5, value=row.qyqudao)
        table.cell(row=i, column=6, value=row.beizhu)
        table.cell(row=i, column=7, value=row.countNumber)
        table.cell(row=i, column=8, value=row.isHaveHaiCang)
        table.cell(row=i, column=9, value=row.categoryName)
        table.cell(row=i, column=10, value=row.companyName)
        table.cell(row=i, column=11, value=row.storeName)

        dianhua,tel,mail,qq = getTelAndMail(row.TelAndMailInfos)
        table.cell(row=i, column=12, value=tel)
        table.cell(row=i, column=13, value=dianhua)
        table.cell(row=i, column=14, value=qq)
        table.cell(row=i, column=15, value=mail)
        table.cell(row=i, column=16, value=row.url)
        table.cell(row=i, column=17, value=row.companyLink)
        temp = row.jystatus
        if temp != None:
            jiyingstatus = row.jystatus.name
        else:
            jiyingstatus = ""
        table.cell(row=i, column=18, value=jiyingstatus) #经营状态
        table.cell(row=i, column=19, value=row.daibiaoren)
        temp = row.zhuceziben
        if temp != None:
            # zhuceziben = row.jystatus.name
            zhuceziben = str(row.zhuceziben) + "万元人民币"
        else:
            zhuceziben = ""

        table.cell(row=i, column=20, value=zhuceziben)  # 注册资本
        table.cell(row=i, column=21, value=row.createtime)
        table.cell(row=i, column=22, value=row.shengfen)  # 所属省份
        table.cell(row=i, column=23, value=row.city)
        table.cell(row=i, column=24, value=row.shehuixinyongma)
        table.cell(row=i, column=25, value=row.nashuishibie)  # 纳税人识别号
        table.cell(row=i, column=26, value=row.yingyezhizhao)
        table.cell(row=i, column=27, value=row.zuzhijigouma)  # 组织机构代码
        table.cell(row=i, column=28, value=row.canbaorenshu)  # 参保人数
        temp = row.qylx
        if temp != None:
            qiyeType = row.qylx.name
        else:
            qiyeType = ""
        table.cell(row=i, column=29, value=qiyeType)  # 企业类型
        table.cell(row=i, column=30, value=row.suoshuhangye)  # 所属行业
        table.cell(row=i, column=31, value=row.registeraddress)
        table.cell(row=i, column=32, value=row.jiyingfanwei)
        table.cell(row=i, column=33, value=row.dengjijiguan)
        table.cell(row=i, column=34, value=row.datafrom)
        table.cell(row=i, column=35, value=row.updatetime)

    if isFromFileNumber==1: #说明是从上次读取的位置读取的，需要更新文件
        writeFile("pythonLastRead.txt",str(lastRowNumber))

    # 调整列宽
    # table.column_dimensions['A'].width = 20.0

    # 生成前14个大写字母  ascii_uppercase生成所有大写字母
    # upper_string = string.ascii_uppercase[:6]
    # for col in upper_string:
    #     table.column_dimensions[col].width = 15
    #
    # # 调整行高
    # table.row_dimensions[1].height = 35


    # 日期
    now_time = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')
    # 格式化时间为中文年月日
    # now_time = datetime.datetime.now().strftime('%Y%m%d %H%M')
    # now_time = now_time[:4] + '年' + now_time[4:6] + '月' + now_time[6:8] + '日' + now_time[8:11] + '时' + now_time[11:13] + '分'
    # now_time = now_time[:4] + '年' + now_time[4:6] + '月' + now_time[6:8] + '日'

    # haveRead = getHaveReadInfo();
    haveRead = "读取了";
    tiaoshu = i-1
    excelName="{}-{}-{}-{}条.xlsx".format(excelName,now_time,haveRead,tiaoshu)
    FilePath = target_path+"/"+excelName
    if(delFile(FilePath)):
        wb.save(FilePath)
        print("保存文件成功:{}，数据条数:{}".format(FilePath,tiaoshu))
        return{"flag":True,"path":excelName}


def delFile(file):

    try:
        # 判断文件是否存在
        if (os.path.exists(file)):
            os.remove(file)
            # print('移除文件：%s' % file)
        # else:
        #     print("无相同文件不用删除！")
        return True
    except Exception as e:
        print('文件{}删除错误，可能文件被打开了，请关闭重试!'.format(file))
        return False


def export(companyName,storeName,userId,bianhao,shengfen,city, \
                          countNumber, canbaorenshu, isHaveCang, startNumber, \
                          endNumber, yingxiaocount, startTime,endTime,isFromFileNumber,target_path):


    readStart = readFile("pythonLastRead.txt")

    if isFromFileNumber==1:# 需要从上次读取的数据库位置开始读取,上次读取的位置存在了pythonLastRead.txt中
        if readStart in ["",'','\n','\r\n']: #文件内容为空
             return {"flag":False,"tips":"请确保PythonLastRead配置文件正确"}
    else:
        readStart = int(readStart)

    file_name = '服务器导出文件'

    results=getData(readStart,companyName,storeName,userId,bianhao,shengfen,city, \
                          countNumber, canbaorenshu, isHaveCang, startNumber, \
                          endNumber, yingxiaocount, startTime,endTime,isFromFileNumber)
    # print("================长度如下===================")
    # print(len(results))
    # return {"flag": True, "tips": "导出文件成功"}
    if(len(results)>0):
        resultFile=generateExcel(results,target_path,file_name,isFromFileNumber)
        print("--> Finished generate excel")
        return {"flag": resultFile["flag"], "tips": "导出文件成功","path":resultFile["path"]}
    else:
        print("没有从数据库中获取数据，可能行号过大")
        return {"flag": True, "tips": "没有从数据库中获取数据，可能行号过大","path":""}



def main(opts):

    readStart = ''
    readStart = readFile("pythonLastRead.txt")
    # print(type(readStart))
    # print(readStart in ["",'','\n','\r\n'])
    if readStart in ["",'','\n','\r\n']:
        if opts.start == -1:
            print("请用-s 作为参数，输入行号起始位置")
            exit()
    if opts.start != -1:
        readStart = opts.start
    if opts.file_name == '':
        print("请用-s 作为参数，输入文件名称")
        exit()

    # exit()

    print(opts.end)
    results=getData(readStart,opts.end)
    if(len(results)>0):
        generateExcel(results,opts.file_name)
        print("--> Finished generate excel")
    else:
        print("没有从数据库中获取数据，可能行号过大")




if __name__ == '__main__':
    opts = parser.parse_args()
    main(opts)