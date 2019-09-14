#coding:utf-8

import argparse
import xlrd


from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import Color, Fill, Font, Alignment, PatternFill, Border, Side
from openpyxl.cell import Cell
import datetime
import string

import os, sys
import random

import pymysql

# 执行命令
# python main.py -s 32 -e 49 -f "22"


parser = argparse.ArgumentParser(description='从数据库中excel')
# General options



parser.add_argument('-s','--start', required = False,default=-1, type=int,
                    help='开始行位置。')

parser.add_argument('-e','--end',required = False,default=-1,type=int, help='结束行位置。')

parser.add_argument('-f','--file_name', required = True,default='', type=str, help='name of the  file.')


def getHaveReadInfo():
    fname = 'haveRead.txt'
    with open(fname, 'r', encoding='utf-8') as f:  # 打开文件
        lines = f.readlines()  # 读取所有行
        if(lines != "" and len(lines)>0):
            first_line = lines[0]  # 取第一行
            # last_line = lines[-1]  # 取最后一行
            print('文件' + fname + '第一行为：' + first_line)
            # print('文件' + fname + '最后一行为：' + last_line)
            return first_line
        else:
            return "已读信息"
def isdigit(aString):
    try:
        x = float(aString)
        return True
    except ValueError as e:
        # print('input is not a number!')
        return False

def getUserInfo():

    user={}

    # 打开数据库连接
    db = pymysql.connect(host='13.231.165.68', port=33070, user='root', passwd='880309jQl', db='tfs', charset='utf8')

    # 使用cursor()方法获取操作游标
    cursor = db.cursor()


    sql = "SELECT * FROM  user"

    try:
        # 执行SQL语句
        cursor.execute(sql)
        # 获取所有记录列表
        results = cursor.fetchall()

        for row in results:
            user[row[0]]=row[5]
        # print(len(results))

    except:
        print("Error: unable to fecth user data")

    # 关闭数据库连接
    db.close()
    return user;


def getData(start,end):
    # 打开数据库连接
    # db = pymysql.connect("localhost", "root", "123456", "tfs", charset='utf8')

    db = pymysql.connect(host='13.231.165.68', port=33070, user='root', passwd='880309jQl', db='tfs', charset='utf8')

    # 使用cursor()方法获取操作游标
    cursor = db.cursor()

    # SQL 查询语句
    if(end!=-1):
        sql = "SELECT * FROM companyinfo  WHERE id BETWEEN {} and {}".format(start, end)
    else:
        sql = "SELECT * FROM companyinfo  WHERE id> {}".format(start)
    try:
        # 执行SQL语句
        cursor.execute(sql)
        # 获取所有记录列表
        results = cursor.fetchall()
        return results
        # print(len(results))

    except:
        print("Error: unable to fecth data")

    # 关闭数据库连接
    db.close()

def readFile(path):
    with open(path, 'r') as my_file:
        # values2 = my_file.readlines()
        # values2 = my_file.read()
        values2 = my_file.readline()
        return values2

def writeFile(path,content):

    with open(path, 'w') as my_file:
        # values2 = my_file.readlines()
        # values2 = my_file.read()
        values2 = my_file.write(content)

def generateExcel(results,excelName):



    # 创建一个工作薄
    wb = Workbook()

    # 创建一个工作表(注意是一个属性) # 激活 worksheet
    table = wb.active

    # excel创建的工作表名默认为sheet1,一下代码实现了给新创建的工作表创建一个新的名字
    table.title = 'ali数据'


    # 设置表头字体居中
    font = Font(name=u'宋体',size=16, bold = True)
    alignment_style = Alignment(horizontal='center', vertical='center')

    # print(type(results))
    i=1
    table.cell(row=i, column=1, value="序号")
    table.cell(row=i, column=2, value="类别名")
    table.cell(row=i, column=3, value="公司名称")
    table.cell(row=i, column=4, value="公司连接")
    table.cell(row=i, column=5, value="商店名")
    table.cell(row=i, column=6, value="商店数量")
    table.cell(row=i, column=7, value="有/无海外仓")
    table.cell(row=i, column=8, value="城市")
    table.cell(row=i, column=9, value="所属人员")
    table.cell(row=i, column=10, value="用户编号")
    table.cell(row=i, column=11, value="社会信用码")
    table.cell(row=i, column=12, value="营业执照")
    table.cell(row=i, column=13, value="注册地址")
    table.cell(row=i, column=14, value="法人代表")
    table.cell(row=i, column=15, value="经营范围")
    table.cell(row=i, column=16, value="注册时间")
    table.cell(row=i, column=17, value="登记机关")
    table.cell(row=i, column=18, value="促销次数")
    table.cell(row=i, column=19, value="备注")
    table.cell(row=i, column=20, value="营销渠道")
    table.cell(row=i, column=21, value="店铺链接(|)")


    lastRowNumber=-1
    for row in results:
        lastRowNumber=row[0]
        i = i + 1
        j = 0
        for ele in row:
            # print(ele)
            j = j+1
            table.cell(row=i, column=j, value=ele)
            # table.cell(row=i, column=2, value=row[1])
            # table.cell(row=i, column=3, value=row[2])
            # table.cell(row=i, column=4, value=row[3])
            # table.cell(row=i, column=5, value=row[4])
            # table.cell(row=i, column=6, value=row[5])
            # table.cell(row=i, column=7, value=row[6])
            # table.cell(row=i, column=8, value=row[7])
            # table.cell(row=i, column=9, value=row[8])
            # table.cell(row=i, column=10, value=row[9])
            # table.cell(row=i, column=11, value=row[10])
            # table.cell(row=i, column=12, value=row[11])
            # table.cell(row=i, column=13, value=row[12])


    writeFile("pythonLastRead.txt",str(lastRowNumber))






    # 调整列宽
    # table.column_dimensions['A'].width = 20.0

    # 生成前14个大写字母  ascii_uppercase生成所有大写字母
    # upper_string = string.ascii_uppercase[:6]
    # for col in upper_string:
    #     table.column_dimensions[col].width = 15
    #
    # # 调整行高
    # table.row_dimensions[1].height = 35


    # 日期
    now_time = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')
    # 格式化时间为中文年月日
    # now_time = datetime.datetime.now().strftime('%Y%m%d %H%M')
    # now_time = now_time[:4] + '年' + now_time[4:6] + '月' + now_time[6:8] + '日' + now_time[8:11] + '时' + now_time[11:13] + '分'
    # now_time = now_time[:4] + '年' + now_time[4:6] + '月' + now_time[6:8] + '日'

    haveRead = getHaveReadInfo();
    tiaoshu = i-1
    excelName="{}-{}-{}-{}条.xlsx".format(excelName,now_time,haveRead,tiaoshu)

    if(delFile(excelName)):
        wb.save(excelName)
        print("保存文件成功:{}，数据条数:{}".format(excelName,tiaoshu))


def delFile(file):

    try:
        # 判断文件是否存在
        if (os.path.exists(file)):
            os.remove(file)
            # print('移除文件：%s' % file)
        # else:
        #     print("无相同文件不用删除！")
        return True
    except Exception as e:
        print('文件{}删除错误，可能文件被打开了，请关闭重试!'.format(file))
        return False




def main(opts):

    readStart = ''
    readStart = readFile("pythonLastRead.txt")
    # print(type(readStart))
    # print(readStart in ["",'','\n','\r\n'])
    if readStart in ["",'','\n','\r\n']:
        if opts.start == -1:
            print("请用-s 作为参数，输入行号起始位置")
            exit()
    if opts.start != -1:
        readStart = opts.start
    if opts.file_name == '':
        print("请用-s 作为参数，输入文件名称")
        exit()

    # exit()

    print(opts.end)
    results=getData(readStart,opts.end)
    if(len(results)>0):
        generateExcel(results,opts.file_name)
        print("--> Finished generate excel")
    else:
        print("没有从数据库中获取数据，可能行号过大")




if __name__ == '__main__':
    opts = parser.parse_args()
    main(opts)